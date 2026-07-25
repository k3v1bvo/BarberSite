import os
import re
import json
import uuid
import datetime
from decimal import Decimal
from openpyxl import load_workbook

# ------------------------------------------------------------------------------
# CONFIGURACIÓN
# ------------------------------------------------------------------------------
XLSX_PATH = "C:/Users/PC/Downloads/migracionesss/REGISTROS DIARIOS CAJA CHICA (ACT 24-05-26).xlsm"
OUT_DIR = os.path.join(os.path.dirname(__file__), "out")

# Espacio de nombres para UUIDs determinísticos
NAMESPACE = uuid.UUID("12345678-1234-5678-1234-567812345678")

# Diccionarios para buscar cabeceras flexibles
SHEETS_CONFIG = {
    "CAJA CHICA": ["FECHA", "C.I.", "NOMBRE", "CUENTA", "DETALLE CUENTA", "GLOSA", "COSTO"],
    "REG. INGRESOS": ["FECHA", "C.I.", "NOMBRE", "CUENTA", "DETALLE CUENTA", "GLOSA", "IVA", "IT", "COSTO", "OPERARIO"],
    "REG. EGRESOS": ["FECHA", "C.I.", "NOMBRE", "CUENTA", "DETALLE CUENTA", "GLOSA", "IVA", "COSTO", "CLASE"],
    "REG. VENTAS": ["FECHA", "C.I.", "NOMBRE", "CUENTA", "DETALLE CUENTA", "GLOSA", "IVA", "IT", "COSTO", "OPERARIO"],
    "REG. CLIENTES": ["C.I.", "NOMBRE"], # Opcional, si existe esta hoja
    "PLAN DE CUENTAS": ["COD", "CUENTA", "NIVEL"], # Asumiendo cabeceras.
}

# ------------------------------------------------------------------------------
# FUNCIONES AUXILIARES
# ------------------------------------------------------------------------------
def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def clean_str(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = re.sub(r"\s+", " ", v.strip())
        return v if v else None
    return str(v).strip()

def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    # Parse manual si es string
    if isinstance(v, str):
        try:
            # Asumimos formato d/m/y o y-m-d
            if "/" in v:
                parts = v.split("/")
                if len(parts) == 3:
                    return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
            elif "-" in v:
                parts = v.split("-")
                if len(parts) == 3:
                    return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
        except:
            pass
    return None

def gen_uuid(*args):
    """Genera un UUID determinístico basado en los argumentos."""
    text = "|".join(str(a) for a in args).encode("utf-8")
    return str(uuid.uuid5(NAMESPACE, text))

def tipo_cuenta(codigo):
    if not codigo: return "EGRESO" # Por defecto
    first = str(codigo).split(".")[0]
    return {
        "1": "ACTIVO",
        "2": "PASIVO",
        "3": "PATRIMONIO",
        "4": "INGRESO",
        "5": "EGRESO",
    }.get(first, "EGRESO")

def metodo_pago(glosa):
    g = (glosa or "").upper()
    if "QR" in g: return "transferencia" # o "qr" según la BD
    if "P.O.S" in g or "POS" in g or "TARJ" in g: return "tarjeta"
    return "efectivo"

def escape_sql(val):
    if val is None:
        return "NULL"
    if isinstance(val, (int, float, Decimal)):
        return str(val)
    if isinstance(val, datetime.date):
        return f"'{val.isoformat()}'"
    # string escape
    s = str(val).replace("'", "''")
    return f"'{s}'"

def format_sql_insert(table, data_dict, on_conflict="DO NOTHING"):
    if not data_dict: return ""
    cols = ", ".join(data_dict.keys())
    vals = ", ".join(escape_sql(v) for v in data_dict.values())
    return f"INSERT INTO {table} ({cols}) VALUES ({vals}) ON CONFLICT {on_conflict};\n"

# ------------------------------------------------------------------------------
# PROCESAMIENTO
# ------------------------------------------------------------------------------
def main():
    ensure_dir(OUT_DIR)
    print(f"Cargando Excel desde: {XLSX_PATH} ...")
    
    if not os.path.exists(XLSX_PATH):
        print("ERROR: El archivo no existe.")
        return

    wb = load_workbook(XLSX_PATH, data_only=True, keep_vba=False)
    print("Excel cargado. Procesando hojas...")

    # Estructuras de datos a recolectar
    cuentas = {} # codigo -> dict
    clientes = {} # ci -> dict
    servicios = {} # codigo_cuenta -> dict
    productos = {} # codigo_cuenta -> dict
    citas = []
    transactions = []
    operarios_set = set()
    
    anomalias = []
    
    def find_header(ws, headers_wanted):
        wanted = [h.upper() for h in headers_wanted]
        for r in range(1, 20):
            row_vals = [clean_str(c.value) or "" for c in ws[r]]
            row_upper = [v.upper() for v in row_vals]
            # check if all wanted are in row
            if all(any(h in cell for cell in row_upper) for h in wanted):
                # map header to col index (1-based)
                cols = []
                for h in wanted:
                    for i, cell in enumerate(row_upper):
                        if h in cell:
                            cols.append(i + 1)
                            break
                return r, cols
        return None, None

    # 1. Extraer Plan de Cuentas si existe
    if "PLAN DE CUENTAS" in wb.sheetnames:
        ws = wb["PLAN DE CUENTAS"]
        for r in range(2, ws.max_row + 1):
            cod = clean_str(ws.cell(r, 1).value)
            det = clean_str(ws.cell(r, 2).value)
            if cod and det:
                cuentas[cod] = {
                    "codigo": cod,
                    "nombre": det,
                    "tipo": tipo_cuenta(cod),
                    "descripcion": ""
                }
    
    # 2. Extraer Clientes (Hoja si existe)
    if "REG. CLIENTES" in wb.sheetnames:
        ws = wb["REG. CLIENTES"]
        hr, cols = find_header(ws, ["C.I.", "NOMBRE"])
        if hr:
            for r in range(hr + 1, ws.max_row + 1):
                ci = clean_str(ws.cell(r, cols[0]).value)
                nom = clean_str(ws.cell(r, cols[1]).value)
                if ci and nom:
                    if ci not in clientes:
                        clientes[ci] = {"id": gen_uuid("cliente", ci), "ci": ci, "nombre": nom}

    # 3. Procesar Movimientos
    movimientos_raw = []
    for sheet in ["CAJA CHICA", "REG. INGRESOS", "REG. EGRESOS", "REG. VENTAS"]:
        if sheet not in wb.sheetnames:
            anomalias.append(f"Hoja '{sheet}' no encontrada en el Excel.")
            continue
            
        ws = wb[sheet]
        headers = SHEETS_CONFIG[sheet]
        hr, cols = find_header(ws, headers)
        
        if not hr:
            anomalias.append(f"Cabeceras no encontradas en '{sheet}'. Esperaba: {headers}")
            continue

        for r in range(hr + 1, ws.max_row + 1):
            rec = {h: clean_str(ws.cell(r, c).value) for h, c in zip(headers, cols)}
            # Extract basic fields
            fecha = as_date(ws.cell(r, cols[headers.index("FECHA")]).value)
            costo = rec.get("COSTO")
            
            # Limpiar costo
            try:
                if costo is not None:
                    costo_str = str(costo).replace('Bs.', '').replace(' ', '').replace(',', '.')
                    costo = float(costo_str)
                else:
                    costo = 0.0
            except:
                costo = 0.0
                
            if not fecha or costo == 0.0:
                continue # Saltar vacíos

            rec["FECHA_CLEAN"] = fecha
            rec["COSTO_CLEAN"] = costo
            rec["SHEET"] = sheet
            rec["ROW"] = r
            movimientos_raw.append(rec)

    # Procesar lógica de negocios sobre movimientos
    for mov in movimientos_raw:
        sheet = mov["SHEET"]
        ci = mov.get("C.I.")
        nombre = mov.get("NOMBRE")
        cuenta_cod = mov.get("CUENTA")
        cuenta_det = mov.get("DETALLE CUENTA")
        glosa = mov.get("GLOSA")
        costo = mov["COSTO_CLEAN"]
        fecha = mov["FECHA_CLEAN"]
        operario = mov.get("OPERARIO")
        
        # Guardar operarios detectados
        if operario:
            operarios_set.add(operario)
            
        # Actualizar clientes faltantes
        if ci and ci.lower() not in ['s/n', 'sn', '0']:
            if ci not in clientes:
                clientes[ci] = {"id": gen_uuid("cliente", ci), "ci": ci, "nombre": nombre or "Desconocido"}
            elif not clientes[ci]["nombre"] and nombre:
                clientes[ci]["nombre"] = nombre
                
        cliente_id = clientes[ci]["id"] if ci in clientes else None

        # Actualizar cuentas faltantes
        if cuenta_cod and cuenta_cod not in cuentas:
            cuentas[cuenta_cod] = {
                "codigo": cuenta_cod,
                "nombre": cuenta_det or "Sin detalle",
                "tipo": tipo_cuenta(cuenta_cod),
                "descripcion": ""
            }
            anomalias.append(f"Cuenta autogenerada desde movimientos: {cuenta_cod} - {cuenta_det}")

        # Identificar Servicios y Productos (heurística basada en cuentas)
        if cuenta_cod and str(cuenta_cod).startswith("4.1.2"): # Asumimos 4.1.2.* son servicios
            if cuenta_cod not in servicios:
                servicios[cuenta_cod] = {
                    "id": gen_uuid("servicio", cuenta_cod),
                    "nombre": cuenta_det or f"Servicio {cuenta_cod}",
                    "precio": costo,
                    "duracion": 30, # Default
                    "cuenta_codigo": cuenta_cod
                }
        elif cuenta_cod and str(cuenta_cod).startswith("4.1.1"): # Asumimos 4.1.1.* son productos
            if cuenta_cod not in productos:
                productos[cuenta_cod] = {
                    "id": gen_uuid("producto", cuenta_cod),
                    "nombre": cuenta_det or f"Producto {cuenta_cod}",
                    "precio": costo,
                    "stock": 0,
                    "cuenta_codigo": cuenta_cod
                }
                
        # Construir Citas (Desde REG. INGRESOS y que sean Servicios)
        if sheet == "REG. INGRESOS" and cuenta_cod and str(cuenta_cod).startswith("4.1.2"):
            cita_id = gen_uuid("cita", sheet, mov["ROW"], fecha, ci)
            citas.append({
                "id": cita_id,
                "cliente_id": cliente_id,
                "servicio_id": servicios[cuenta_cod]["id"],
                "fecha_hora": f"{fecha} 12:00:00", # Asumimos medio día
                "estado": "completado",
                "notas": f"Importado Excel. Op: {operario}. Glosa: {glosa}",
                "precio_final": costo,
                "metodo_pago": metodo_pago(glosa)
            })

        # Construir Transactions
        # Mapeo:
        # CAJA CHICA     -> libro = CAJA_CHICA, tipo = INGRESO/EGRESO
        # REG. INGRESOS  -> libro = SERVICIOS,   tipo = INGRESO
        # REG. VENTAS    -> libro = VENTAS,      tipo = INGRESO
        # REG. EGRESOS   -> libro = CAJA_CHICA,  tipo = EGRESO
        
        if sheet == "CAJA CHICA":
            libro = "CAJA_CHICA"
            tipo_mov = "INGRESO" if tipo_cuenta(cuenta_cod) == "INGRESO" else "EGRESO"
        elif sheet == "REG. INGRESOS":
            libro = "SERVICIOS"
            tipo_mov = "INGRESO"
        elif sheet == "REG. VENTAS":
            libro = "VENTAS"
            tipo_mov = "INGRESO"
        elif sheet == "REG. EGRESOS":
            libro = "CAJA_CHICA"
            tipo_mov = "EGRESO"
        else:
            libro = "CAJA_CHICA"
            tipo_mov = "EGRESO"

        trans_id = gen_uuid("trans", sheet, mov["ROW"], fecha, costo)
        
        # Construimos el texto para "notas" para poder enlazar luego al empleado
        notas_arr = []
        if operario: notas_arr.append(f"Op: {operario}")
        if glosa: notas_arr.append(f"Glosa: {glosa}")
        if sheet: notas_arr.append(f"Hoja: {sheet} Fila: {mov['ROW']}")
        
        transactions.append({
            "id": trans_id,
            "libro": libro,
            "tipo_movimiento": tipo_mov,
            "fecha": fecha,
            "monto": costo,
            "metodo_pago": metodo_pago(glosa),
            "cuenta_codigo": cuenta_cod,
            "cliente_id": cliente_id,
            "descripcion": cuenta_det or "Sin detalle",
            "notas": " | ".join(notas_arr)
        })

    # ------------------------------------------------------------------------------
    # GENERAR ARCHIVOS SQL
    # ------------------------------------------------------------------------------
    
    # 1. PLAN DE CUENTAS
    with open(os.path.join(OUT_DIR, "01_plan_cuentas.sql"), "w", encoding="utf-8") as f:
        f.write("-- GENERADO AUTOMÁTICAMENTE\n")
        f.write("BEGIN;\n")
        for c in cuentas.values():
            f.write(format_sql_insert("public.plan_cuentas", {
                "codigo": c["codigo"],
                "nombre": c["nombre"],
                "tipo": c["tipo"],
                "descripcion": c["descripcion"]
            }, "(codigo) DO NOTHING"))
        f.write("COMMIT;\n")

    # 2. CLIENTES
    with open(os.path.join(OUT_DIR, "02_clientes.sql"), "w", encoding="utf-8") as f:
        f.write("-- GENERADO AUTOMÁTICAMENTE\n")
        f.write("BEGIN;\n")
        for c in clientes.values():
            f.write(format_sql_insert("public.clientes", {
                "id": c["id"],
                "ci": c["ci"],
                "nombre": c["nombre"]
            }, "(ci) DO NOTHING"))
        f.write("COMMIT;\n")

    # 3. SERVICIOS
    with open(os.path.join(OUT_DIR, "03_servicios.sql"), "w", encoding="utf-8") as f:
        f.write("-- GENERADO AUTOMÁTICAMENTE\n")
        f.write("BEGIN;\n")
        for s in servicios.values():
            f.write(format_sql_insert("public.servicios", {
                "id": s["id"],
                "nombre": s["nombre"],
                "precio": s["precio"],
                "duracion": s["duracion"],
                "cuenta_codigo": s["cuenta_codigo"]
            }, "(id) DO NOTHING"))
        f.write("COMMIT;\n")

    # 4. PRODUCTOS
    with open(os.path.join(OUT_DIR, "04_productos.sql"), "w", encoding="utf-8") as f:
        f.write("-- GENERADO AUTOMÁTICAMENTE\n")
        f.write("BEGIN;\n")
        for p in productos.values():
            f.write(format_sql_insert("public.productos", {
                "id": p["id"],
                "nombre": p["nombre"],
                "precio": p["precio"],
                "stock": p["stock"],
                "cuenta_codigo": p["cuenta_codigo"]
            }, "(id) DO NOTHING"))
        f.write("COMMIT;\n")

    # 5. CITAS
    with open(os.path.join(OUT_DIR, "05_citas.sql"), "w", encoding="utf-8") as f:
        f.write("-- GENERADO AUTOMÁTICAMENTE\n")
        f.write("BEGIN;\n")
        for c in citas:
            f.write(format_sql_insert("public.citas", c, "(id) DO NOTHING"))
        f.write("COMMIT;\n")

    # 6. TRANSACTIONS
    with open(os.path.join(OUT_DIR, "06_transactions.sql"), "w", encoding="utf-8") as f:
        f.write("-- GENERADO AUTOMÁTICAMENTE\n")
        f.write("BEGIN;\n")
        for t in transactions:
            f.write(format_sql_insert("public.transactions", t, "(id) DO NOTHING"))
        f.write("COMMIT;\n")

    # ALIASES
    with open(os.path.join(OUT_DIR, "operarios_aliases.csv"), "w", encoding="utf-8") as f:
        f.write("ALIAS_EXCEL,UUID_PERFIL_FINAL\n")
        for op in sorted(list(operarios_set)):
            f.write(f'"{op}",\n')

    # REPORTE
    with open(os.path.join(OUT_DIR, "reporte_migracion.md"), "w", encoding="utf-8") as f:
        f.write("# Reporte de Extracción de Datos Excel\n\n")
        f.write(f"**Cuentas encontradas:** {len(cuentas)}\n")
        f.write(f"**Clientes únicos:** {len(clientes)}\n")
        f.write(f"**Servicios identificados:** {len(servicios)}\n")
        f.write(f"**Productos identificados:** {len(productos)}\n")
        f.write(f"**Citas generadas:** {len(citas)}\n")
        f.write(f"**Transacciones (Movimientos) generadas:** {len(transactions)}\n\n")
        f.write("## Operarios (Requieren creación manual y mapeo)\n")
        for op in sorted(list(operarios_set)):
            f.write(f"- {op}\n")
        f.write("\n## Anomalías / Alertas\n")
        if not anomalias:
            f.write("No se detectaron anomalías mayores.\n")
        else:
            for a in set(anomalias):
                f.write(f"- {a}\n")

    print(f"✅ Extracción completada. Archivos guardados en {OUT_DIR}")

if __name__ == "__main__":
    main()
